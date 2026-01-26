#!/usr/bin/env python3
"""
Functional Annotation Pipeline - Excel Output Generator

Description:
    This script parses output files from functional annotation tools
    (KofamScan, InterProScan, EggNOG-mapper, and FANTASIA) and generates Excel
    files with standardized format. For each tool, the script generates two types of outputs:
    
    1. Per-term output: One row per functional term (GO or KEGG)
    2. Per-gene output: One row per gene with functional terms grouped together
    
    KofamScan:
        - Per-term: gene_name, KEGG
        - Per-gene: gene_name, KEGG (grouped)
    
    InterProScan:
        - Per-term: gene, analysis, score, InterPro_accession, InterPro_description, GO, Pathways
        - Per-gene: gene, GO (grouped), Pathways (grouped)
    
    EggNOG-mapper v5:
        - Per-term: gene, term_type, term
        - Per-gene: gene, Description, GOs, KEGG_ko, KEGG_Pathway, KEGG_Reaction, KEGG_rclass, PFAM
    
    EggNOG-mapper v7:
        - Per-term: gene, term_type, term, score
        - Per-gene: gene, eggnog_protein_ID, GOs (with scores), KEGGs (with scores)
    
    FANTASIA:
        - Per-term: gene, GO, term_count, final_score (separate file per model)
        - Per-gene: gene, GO (grouped) (separate file per model)

Author: Aaron Kiggen
Repository: aaronkiggen/Functional_Annotation_Pipeline
"""

import argparse
import gzip
import os
import sys
from pathlib import Path
from typing import List, Dict, Tuple

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Error: Required packages not found.")
    print("Please install: pip install pandas openpyxl")
    sys.exit(1)


class AnnotationParser:
    """Base class for parsing annotation files"""
    
    # Constants
    KEGG_PREFIX = 'ko:'
    KEGG_PREFIX_LEN = len(KEGG_PREFIX)
    
    def __init__(self, input_file: str):
        self.input_file = input_file
        self.results = []
    
    def parse(self):
        """Parse the input file - to be implemented by subclasses"""
        raise NotImplementedError
    
    def to_dataframe(self) -> pd.DataFrame:
        """Convert results to pandas DataFrame"""
        if not self.results:
            return pd.DataFrame(columns=['gene_name', 'functional_term', 'extra_information'])
        return pd.DataFrame(self.results)
    
    def save_to_excel(self, output_file: str):
        """Save results to Excel file with formatting"""
        df = self.to_dataframe()
        
        if df.empty:
            print(f"Warning: No data to write for {output_file}")
            return
        
        # Create Excel writer
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Annotations')
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Annotations']
            
            # Format header
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Adjust column widths
            worksheet.column_dimensions['A'].width = 25  # gene_name
            worksheet.column_dimensions['B'].width = 20  # functional_term
            worksheet.column_dimensions['C'].width = 60  # extra_information
        
        print(f"✓ Created: {output_file} ({len(df)} rows)")


class KofamScanParser(AnnotationParser):
    """Parser for KofamScan output files"""
    
    def parse(self):
        """
        Parse KofamScan pre-script output format:
        gene_name
        gene_name    KEGG_term
        """
        self.results = []
        
        if not os.path.exists(self.input_file):
            print(f"Warning: File not found: {self.input_file}")
            return
        
        with open(self.input_file, 'r') as f:
            for line in f:
                # Skip comments and empty lines
                if line.startswith('#') or not line.strip():
                    continue
                
                parts = line.strip().split('\t')
                gene_name = parts[0] if len(parts) >= 1 else ""
                
                # Gene may or may not have a KEGG term
                if len(parts) >= 2:
                    kegg_term = parts[1]
                    self.results.append({
                        'gene_name': gene_name,
                        'KEGG': kegg_term
                    })
                else:
                    # Gene without KEGG term
                    self.results.append({
                        'gene_name': gene_name,
                        'KEGG': ''
                    })
    
    def to_dataframe(self) -> pd.DataFrame:
        """Convert results to pandas DataFrame with KofamScan-specific columns"""
        if not self.results:
            return pd.DataFrame(columns=['gene_name', 'KEGG'])
        return pd.DataFrame(self.results)
    
    def save_to_excel(self, output_file: str):
        """Save results to Excel file with KofamScan-specific formatting"""
        df = self.to_dataframe()
        
        if df.empty:
            print(f"Warning: No data to write for {output_file}")
            return
        
        # Create Excel writer
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Annotations')
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Annotations']
            
            # Format header
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Adjust column widths
            worksheet.column_dimensions['A'].width = 25  # gene_name
            worksheet.column_dimensions['B'].width = 20  # KEGG
        
        print(f"✓ Created: {output_file} ({len(df)} rows)")
    
    def create_per_gene_output(self) -> pd.DataFrame:
        """Create per-gene output with grouped KEGG terms"""
        if not self.results:
            return pd.DataFrame(columns=['gene_name', 'KEGG'])
        
        df = pd.DataFrame(self.results)
        # Group by gene and concatenate KEGG terms
        per_gene = df.groupby('gene_name')['KEGG'].apply(lambda x: ', '.join(filter(None, x))).reset_index()
        return per_gene


class InterProScanParser(AnnotationParser):
    """Parser for InterProScan TSV output files"""
    
    def parse(self):
        """
        Parse InterProScan TSV format:
        Desired output: gene / analysis (col 4) / score (col 9) / col 12 / col 13 / GO (col 14) / Pathways (col 15)
        TSV columns: 0=Protein accession, 3=Analysis, 8=Score, 11=InterPro accession, 
                    12=InterPro description, 13=GO annotations, 14=Pathways
        """
        self.results = []
        
        if not os.path.exists(self.input_file):
            print(f"Warning: File not found: {self.input_file}")
            return
        
        with open(self.input_file, 'r') as f:
            for line in f:
                # Skip comments
                if line.startswith('#'):
                    continue
                
                parts = line.strip().split('\t')
                if len(parts) < 4:
                    continue
                
                gene_name = parts[0] if len(parts) > 0 else ""
                analysis = parts[3] if len(parts) > 3 else ""  # Column 4 (0-indexed = 3)
                score = parts[8] if len(parts) > 8 else ""     # Column 9 (0-indexed = 8)
                interpro_acc = parts[11] if len(parts) > 11 else ""  # Column 12 (0-indexed = 11)
                interpro_desc = parts[12] if len(parts) > 12 else ""  # Column 13 (0-indexed = 12)
                go_terms = parts[13] if len(parts) > 13 else ""      # Column 14 (0-indexed = 13)
                pathways = parts[14] if len(parts) > 14 else ""      # Column 15 (0-indexed = 14)
                
                # Store one row per hit
                self.results.append({
                    'gene': gene_name,
                    'analysis': analysis,
                    'score': score,
                    'InterPro_accession': interpro_acc,
                    'InterPro_description': interpro_desc,
                    'GO': go_terms,
                    'Pathways': pathways
                })
    
    def to_dataframe(self) -> pd.DataFrame:
        """Convert results to pandas DataFrame with InterProScan-specific columns"""
        if not self.results:
            return pd.DataFrame(columns=['gene', 'analysis', 'score', 'InterPro_accession', 
                                        'InterPro_description', 'GO', 'Pathways'])
        return pd.DataFrame(self.results)
    
    def save_to_excel(self, output_file: str):
        """Save results to Excel file with InterProScan-specific formatting"""
        df = self.to_dataframe()
        
        if df.empty:
            print(f"Warning: No data to write for {output_file}")
            return
        
        # Create Excel writer
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Annotations')
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Annotations']
            
            # Format header
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Adjust column widths
            worksheet.column_dimensions['A'].width = 25  # gene
            worksheet.column_dimensions['B'].width = 20  # analysis
            worksheet.column_dimensions['C'].width = 15  # score
            worksheet.column_dimensions['D'].width = 20  # InterPro_accession
            worksheet.column_dimensions['E'].width = 40  # InterPro_description
            worksheet.column_dimensions['F'].width = 60  # GO
            worksheet.column_dimensions['G'].width = 30  # Pathways
        
        print(f"✓ Created: {output_file} ({len(df)} rows)")
    
    def create_per_gene_output(self) -> pd.DataFrame:
        """Create per-gene output with grouped GO and Pathways"""
        if not self.results:
            return pd.DataFrame(columns=['gene', 'GO', 'Pathways'])
        
        df = pd.DataFrame(self.results)
        
        # Aggregate GO terms and Pathways per gene
        def aggregate_values(series):
            # Filter out empty strings and join unique values
            values = [v for v in series if v and v.strip()]
            unique_values = []
            for v in values:
                # Split by pipe and comma to handle multiple GO terms in one cell
                sub_values = v.replace('|', ',').split(',')
                for sv in sub_values:
                    sv = sv.strip()
                    if sv and sv not in unique_values:
                        unique_values.append(sv)
            return ', '.join(unique_values)
        
        per_gene = df.groupby('gene').agg({
            'GO': aggregate_values,
            'Pathways': aggregate_values
        }).reset_index()
        
        return per_gene


class EggNOGParser(AnnotationParser):
    """Parser for EggNOG-mapper annotation files"""
    
    def parse(self):
        """
        Parse EggNOG .emapper.annotations format:
        Desired output: gene / Description / GOs / KEGG_ko / KEGG_Pathway / KEGG_Reaction / KEGG_rclass / PFAM
        First 2 rows are comment lines starting with ##
        """
        self.results = []
        
        if not os.path.exists(self.input_file):
            print(f"Warning: File not found: {self.input_file}")
            return
        
        # Determine if file is gzipped
        is_gzipped = self.input_file.endswith('.gz')
        
        open_func = gzip.open if is_gzipped else open
        mode = 'rt' if is_gzipped else 'r'
        
        with open_func(self.input_file, mode) as f:
            header_line = None
            for line in f:
                # Skip comment lines (including the first 2 rows with ##)
                if line.startswith('#'):
                    # Find header line (starts with #query)
                    if line.startswith('#query'):
                        header_line = line.strip().lstrip('#').split('\t')
                    continue
                
                parts = line.strip().split('\t')
                
                if not header_line or len(parts) != len(header_line):
                    continue
                
                # Create dictionary from header and values
                row_dict = dict(zip(header_line, parts))
                
                gene_name = row_dict.get('query', '')
                description = row_dict.get('Description', '') or row_dict.get('eggNOG_desc', '')
                go_terms = row_dict.get('GOs', '') or row_dict.get('GO_terms', '')
                kegg_ko = row_dict.get('KEGG_ko', '') or row_dict.get('KEGG_KO', '')
                kegg_pathway = row_dict.get('KEGG_Pathway', '') or row_dict.get('KEGG_pathway', '')
                kegg_reaction = row_dict.get('KEGG_Reaction', '') or row_dict.get('KEGG_reaction', '')
                kegg_rclass = row_dict.get('KEGG_rclass', '')
                pfam = row_dict.get('PFAMs', '') or row_dict.get('PFAM', '')
                
                # Store one row per gene
                self.results.append({
                    'gene': gene_name,
                    'Description': description if description != '-' else '',
                    'GOs': go_terms if go_terms != '-' else '',
                    'KEGG_ko': kegg_ko if kegg_ko != '-' else '',
                    'KEGG_Pathway': kegg_pathway if kegg_pathway != '-' else '',
                    'KEGG_Reaction': kegg_reaction if kegg_reaction != '-' else '',
                    'KEGG_rclass': kegg_rclass if kegg_rclass != '-' else '',
                    'PFAM': pfam if pfam != '-' else ''
                })
    
    def to_dataframe(self) -> pd.DataFrame:
        """Convert results to pandas DataFrame with EggNOG-specific columns"""
        if not self.results:
            return pd.DataFrame(columns=['gene', 'Description', 'GOs', 'KEGG_ko', 
                                        'KEGG_Pathway', 'KEGG_Reaction', 'KEGG_rclass', 'PFAM'])
        return pd.DataFrame(self.results)
    
    def save_to_excel(self, output_file: str):
        """Save results to Excel file with EggNOG-specific formatting"""
        df = self.to_dataframe()
        
        if df.empty:
            print(f"Warning: No data to write for {output_file}")
            return
        
        # Create Excel writer
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Annotations')
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Annotations']
            
            # Format header
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Adjust column widths
            worksheet.column_dimensions['A'].width = 25  # gene
            worksheet.column_dimensions['B'].width = 50  # Description
            worksheet.column_dimensions['C'].width = 60  # GOs
            worksheet.column_dimensions['D'].width = 20  # KEGG_ko
            worksheet.column_dimensions['E'].width = 30  # KEGG_Pathway
            worksheet.column_dimensions['F'].width = 30  # KEGG_Reaction
            worksheet.column_dimensions['G'].width = 30  # KEGG_rclass
            worksheet.column_dimensions['H'].width = 40  # PFAM
        
        print(f"✓ Created: {output_file} ({len(df)} rows)")
    
    def create_per_term_output(self) -> pd.DataFrame:
        """Create per-term output with 1 row per GO or KEGG term"""
        if not self.results:
            return pd.DataFrame(columns=['gene', 'term_type', 'term'])
        
        rows = []
        for result in self.results:
            gene = result['gene']
            
            # Process GO terms
            if result['GOs']:
                go_list = result['GOs'].split(',')
                for go_term in go_list:
                    go_term = go_term.strip()
                    if go_term:
                        rows.append({
                            'gene': gene,
                            'term_type': 'GO',
                            'term': go_term
                        })
            
            # Process KEGG terms
            if result['KEGG_ko']:
                kegg_list = result['KEGG_ko'].split(',')
                for kegg_term in kegg_list:
                    kegg_term = kegg_term.strip()
                    # Remove 'ko:' prefix if present
                    if kegg_term.startswith('ko:'):
                        kegg_term = kegg_term[3:]
                    if kegg_term:
                        rows.append({
                            'gene': gene,
                            'term_type': 'KEGG',
                            'term': kegg_term
                        })
        
        return pd.DataFrame(rows)


class EggNOG7Parser(AnnotationParser):
    """Parser for EggNOG 7 annotator output files"""
    
    def parse(self):
        """
        Parse EggNOG 7 .eggnog.tsv.gz format with scores:
        Expected format: Column 0 = gene_name, Column 1 = eggnog_protein_ID,
                        Column 12 = KEGG with scores, Column 14 = GO with scores
        Desired output: gene / eggnog_protein_ID / GOs (with scores) / KEGGs (with scores)
        """
        self.results = []
        
        if not os.path.exists(self.input_file):
            print(f"Warning: File not found: {self.input_file}")
            return
        
        # File should be gzipped
        with gzip.open(self.input_file, 'rt') as f:
            header_line = None
            for line in f:
                # Skip comment lines
                if line.startswith('#'):
                    # Find header line (starts with #query)
                    if line.startswith('#query'):
                        header_line = line.strip().lstrip('#').split('\t')
                    continue
                
                parts = line.strip().split('\t')
                
                if not parts or len(parts) < 2:
                    continue
                
                # Use positional columns if we have enough columns
                # Expected: col 0 = gene, col 1 = eggnog_protein_ID
                # col 12 = KEGG with scores, col 14 = GO with scores
                if len(parts) >= 15:
                    gene_name = parts[0]
                    eggnog_protein_id = parts[1]
                    kegg_with_scores = parts[12] if len(parts) > 12 else ''
                    go_with_scores = parts[14] if len(parts) > 14 else ''
                elif header_line and len(parts) == len(header_line):
                    # Fallback to header-based parsing for compatibility
                    row_dict = dict(zip(header_line, parts))
                    gene_name = row_dict.get('query', '')
                    eggnog_protein_id = row_dict.get('seed_ortholog', '')
                    kegg_with_scores = row_dict.get('KEGG_ko', '') or row_dict.get('KEGG_KO', '')
                    go_with_scores = row_dict.get('GOs', '') or row_dict.get('GO', '')
                else:
                    # Not enough columns or no header
                    continue
                
                # Clean up empty values (-, empty strings)
                if eggnog_protein_id == '-':
                    eggnog_protein_id = ''
                if kegg_with_scores == '-':
                    kegg_with_scores = ''
                if go_with_scores == '-':
                    go_with_scores = ''
                
                # Store one row per gene with scores included
                self.results.append({
                    'gene': gene_name,
                    'eggnog_protein_ID': eggnog_protein_id,
                    'GOs': go_with_scores,
                    'KEGGs': kegg_with_scores
                })
    
    
    def to_dataframe(self) -> pd.DataFrame:
        """Convert results to pandas DataFrame with EggNOG v7 columns"""
        if not self.results:
            return pd.DataFrame(columns=['gene', 'eggnog_protein_ID', 'GOs', 'KEGGs'])
        return pd.DataFrame(self.results)
    
    def save_to_excel(self, output_file: str):
        """Save results to Excel file with EggNOG v7 formatting"""
        df = self.to_dataframe()
        
        if df.empty:
            print(f"Warning: No data to write for {output_file}")
            return
        
        # Create Excel writer
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Annotations')
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Annotations']
            
            # Format header
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Adjust column widths
            worksheet.column_dimensions['A'].width = 25  # gene
            worksheet.column_dimensions['B'].width = 30  # eggnog_protein_ID
            worksheet.column_dimensions['C'].width = 80  # GOs
            worksheet.column_dimensions['D'].width = 40  # KEGGs
        
        print(f"✓ Created: {output_file} ({len(df)} rows)")
    
    def create_per_term_output(self) -> pd.DataFrame:
        """Create per-term output with 1 row per GO or KEGG term with scores"""
        if not self.results:
            return pd.DataFrame(columns=['gene', 'term_type', 'term', 'score'])
        
        rows = []
        for result in self.results:
            gene = result['gene']
            
            # Process GO terms with scores (format: GO:0030154|2.33;GO:0048856|2.33)
            if result['GOs']:
                go_items = result['GOs'].split(';')
                for go_item in go_items:
                    go_item = go_item.strip()
                    if go_item and '|' in go_item:
                        # Split term and score
                        term, score = go_item.split('|', 1)
                        rows.append({
                            'gene': gene,
                            'term_type': 'GO',
                            'term': term.strip(),
                            'score': score.strip()
                        })
                    elif go_item:
                        # Handle case without score
                        rows.append({
                            'gene': gene,
                            'term_type': 'GO',
                            'term': go_item,
                            'score': ''
                        })
            
            # Process KEGG terms with scores (format: K25226|46.22;K00001|50.00)
            if result['KEGGs']:
                kegg_items = result['KEGGs'].split(';')
                for kegg_item in kegg_items:
                    kegg_item = kegg_item.strip()
                    if kegg_item and '|' in kegg_item:
                        # Split term and score
                        term, score = kegg_item.split('|', 1)
                        # Remove 'ko:' prefix if present
                        term = term.strip()
                        if term.startswith('ko:'):
                            term = term[3:]
                        rows.append({
                            'gene': gene,
                            'term_type': 'KEGG',
                            'term': term,
                            'score': score.strip()
                        })
                    elif kegg_item:
                        # Handle case without score
                        kegg_item_clean = kegg_item.strip()
                        if kegg_item_clean.startswith('ko:'):
                            kegg_item_clean = kegg_item_clean[3:]
                        rows.append({
                            'gene': gene,
                            'term_type': 'KEGG',
                            'term': kegg_item_clean,
                            'score': ''
                        })
        
        return pd.DataFrame(rows)



class FantasiaParser(AnnotationParser):
    """Parser for FANTASIA AI-driven annotation output files"""
    
    # Model names mapping for output file naming
    MODEL_NAMES = {
        'ESM_L0': 'ESM-2',
        'Prot-T5_L0': 'ProtT5',
        'Prost-T5_L0': 'ProstT5',
        'Ankh3-Large_L0': 'Ankh3-Large',
        'ESM3c_L0': 'ESM3c'
    }
    
    def __init__(self, input_file: str, model_suffix: str = None):
        """
        Initialize parser with optional model suffix to filter specific model
        
        Args:
            input_file: Path to FANTASIA output TSV file
            model_suffix: If specified, only parse data for this model (e.g., 'ESM_L0')
        """
        super().__init__(input_file)
        self.model_suffix = model_suffix
    
    def parse(self):
        """
        Parse FANTASIA TSV format:
        Desired output: gene / GO / term_count / final_score
        Columns: accession, go_id, term_count, final_score_<model>, proteins
        """
        self.results = []
        
        if not os.path.exists(self.input_file):
            print(f"Warning: File not found: {self.input_file}")
            return
        
        with open(self.input_file, 'r') as f:
            header_line = None
            for line in f:
                # Skip comments
                if line.startswith('#'):
                    continue
                
                # First non-comment line is header
                if header_line is None:
                    header_line = line.strip().split('\t')
                    # Validate required columns exist
                    required_cols = ['accession', 'go_id']
                    if not all(col in header_line for col in required_cols):
                        print(f"Warning: Missing required columns in {self.input_file}")
                        print(f"Expected: {required_cols}")
                        return
                    continue
                
                parts = line.strip().split('\t')
                
                # Skip lines that don't match header length
                if len(parts) != len(header_line):
                    continue
                
                # Create dictionary from header and values
                row_dict = dict(zip(header_line, parts))
                
                gene_name = row_dict.get('accession', '')
                go_id = row_dict.get('go_id', '')
                term_count = row_dict.get('term_count', '')
                
                # Skip if no GO term
                if not go_id or go_id == '-':
                    continue
                
                # Get final_score for the specific model
                score_col = f'final_score_{self.model_suffix}' if self.model_suffix else None
                final_score = row_dict.get(score_col, '') if score_col else ''
                
                # Skip if this model has no score for this entry
                if self.model_suffix and (not final_score or final_score.strip() == ''):
                    continue
                
                self.results.append({
                    'gene': gene_name,
                    'GO': go_id,
                    'term_count': term_count,
                    'final_score': final_score
                })
    
    def to_dataframe(self) -> pd.DataFrame:
        """Convert results to pandas DataFrame with FANTASIA-specific column names"""
        if not self.results:
            return pd.DataFrame(columns=['gene', 'GO', 'term_count', 'final_score'])
        return pd.DataFrame(self.results)
    
    def save_to_excel(self, output_file: str):
        """Save results to Excel file with FANTASIA-specific formatting"""
        df = self.to_dataframe()
        
        if df.empty:
            print(f"Warning: No data to write for {output_file}")
            return
        
        # Create Excel writer
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Annotations')
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Annotations']
            
            # Format header
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Adjust column widths for FANTASIA format
            worksheet.column_dimensions['A'].width = 25  # gene
            worksheet.column_dimensions['B'].width = 20  # GO
            worksheet.column_dimensions['C'].width = 15  # term_count
            worksheet.column_dimensions['D'].width = 15  # final_score
        
        print(f"✓ Created: {output_file} ({len(df)} rows)")
    
    def create_per_gene_output(self) -> pd.DataFrame:
        """Create per-gene output with grouped GO terms"""
        if not self.results:
            return pd.DataFrame(columns=['gene', 'GO'])
        
        df = pd.DataFrame(self.results)
        # Group by gene and concatenate GO terms
        per_gene = df.groupby('gene')['GO'].apply(lambda x: ', '.join(x)).reset_index()
        return per_gene


def find_files(directory: str, pattern: str) -> List[str]:
    """Find files matching a pattern in a directory"""
    path = Path(directory)
    if not path.exists():
        return []
    
    files = []
    for file_path in path.glob(pattern):
        if file_path.is_file():
            files.append(str(file_path))
    
    return sorted(files)


def process_kofamscan(results_dir: str, output_dir: str):
    """Process all KofamScan output files"""
    print("\n" + "="*60)
    print("Processing KofamScan Results")
    print("="*60)
    
    kofam_dir = os.path.join(results_dir, 'kofamscan')
    if not os.path.exists(kofam_dir):
        print(f"Warning: KofamScan directory not found: {kofam_dir}")
        return
    
    # Find all mapper files
    mapper_files = find_files(kofam_dir, '*_kofam_mapper.tsv')
    
    if not mapper_files:
        print("No KofamScan mapper files found")
        return
    
    print(f"Found {len(mapper_files)} KofamScan file(s)")
    
    for mapper_file in mapper_files:
        basename = os.path.basename(mapper_file).replace('_kofam_mapper.tsv', '')
        
        parser = KofamScanParser(mapper_file)
        parser.parse()
        
        # Save per-term output (1 row = 1 KEGG term)
        output_file_per_term = os.path.join(output_dir, f"{basename}_kofamscan_per_term.xlsx")
        parser.save_to_excel(output_file_per_term)
        
        # Save per-gene output (1 row = 1 gene with grouped KEGG terms)
        per_gene_df = parser.create_per_gene_output()
        if not per_gene_df.empty:
            output_file_per_gene = os.path.join(output_dir, f"{basename}_kofamscan_per_gene.xlsx")
            with pd.ExcelWriter(output_file_per_gene, engine='openpyxl') as writer:
                per_gene_df.to_excel(writer, index=False, sheet_name='Annotations')
                
                # Get workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Annotations']
                
                # Format header
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
                
                # Adjust column widths
                worksheet.column_dimensions['A'].width = 25  # gene_name
                worksheet.column_dimensions['B'].width = 60  # KEGG (grouped)
            
            print(f"✓ Created: {output_file_per_gene} ({len(per_gene_df)} rows)")


def process_interproscan(results_dir: str, output_dir: str):
    """Process all InterProScan output files"""
    print("\n" + "="*60)
    print("Processing InterProScan Results")
    print("="*60)
    
    interpro_dir = os.path.join(results_dir, 'interproscan')
    if not os.path.exists(interpro_dir):
        print(f"Warning: InterProScan directory not found: {interpro_dir}")
        return
    
    # Find all TSV files
    tsv_files = find_files(interpro_dir, '*.tsv')
    
    if not tsv_files:
        print("No InterProScan TSV files found")
        return
    
    print(f"Found {len(tsv_files)} InterProScan file(s)")
    
    for tsv_file in tsv_files:
        basename = os.path.basename(tsv_file).replace('.tsv', '')
        
        parser = InterProScanParser(tsv_file)
        parser.parse()
        
        # Save per-term output (1 row = 1 hit)
        output_file_per_term = os.path.join(output_dir, f"{basename}_interproscan_per_term.xlsx")
        parser.save_to_excel(output_file_per_term)
        
        # Save per-gene output (1 row = 1 gene with grouped GO and Pathways)
        per_gene_df = parser.create_per_gene_output()
        if not per_gene_df.empty:
            output_file_per_gene = os.path.join(output_dir, f"{basename}_interproscan_per_gene.xlsx")
            with pd.ExcelWriter(output_file_per_gene, engine='openpyxl') as writer:
                per_gene_df.to_excel(writer, index=False, sheet_name='Annotations')
                
                # Get workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Annotations']
                
                # Format header
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
                
                # Adjust column widths
                worksheet.column_dimensions['A'].width = 25  # gene
                worksheet.column_dimensions['B'].width = 80  # GO (grouped)
                worksheet.column_dimensions['C'].width = 40  # Pathways (grouped)
            
            print(f"✓ Created: {output_file_per_gene} ({len(per_gene_df)} rows)")


def process_eggnog(results_dir: str, output_dir: str):
    """Process all EggNOG-mapper output files"""
    print("\n" + "="*60)
    print("Processing EggNOG-mapper Results")
    print("="*60)
    
    # Check for both v5 and v7 outputs
    eggnog_v5_dir = os.path.join(results_dir, 'eggnog', 'v5')
    eggnog_v7_dir = os.path.join(results_dir, 'eggnog7')
    
    # Process v5 outputs
    if os.path.exists(eggnog_v5_dir):
        print("\nProcessing EggNOG v5 outputs...")
        # Find annotation files in subdirectories
        annotation_files = []
        for root, dirs, files in os.walk(eggnog_v5_dir):
            for file in files:
                if file.endswith('.emapper.annotations'):
                    annotation_files.append(os.path.join(root, file))
        
        if annotation_files:
            print(f"Found {len(annotation_files)} EggNOG v5 annotation file(s)")
            for annot_file in annotation_files:
                # Extract sample name from path
                basename = os.path.basename(annot_file).replace('.emapper.annotations', '')
                
                parser = EggNOGParser(annot_file)
                parser.parse()
                
                # Save per-gene output (1 row = 1 gene, already the desired format)
                output_file_per_gene = os.path.join(output_dir, f"{basename}_eggnog_v5_per_gene.xlsx")
                parser.save_to_excel(output_file_per_gene)
                
                # Save per-term output (1 row = 1 GO or KEGG term)
                per_term_df = parser.create_per_term_output()
                if not per_term_df.empty:
                    output_file_per_term = os.path.join(output_dir, f"{basename}_eggnog_v5_per_term.xlsx")
                    with pd.ExcelWriter(output_file_per_term, engine='openpyxl') as writer:
                        per_term_df.to_excel(writer, index=False, sheet_name='Annotations')
                        
                        # Get workbook and worksheet
                        workbook = writer.book
                        worksheet = writer.sheets['Annotations']
                        
                        # Format header
                        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        header_font = Font(bold=True, color="FFFFFF")
                        
                        for cell in worksheet[1]:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal='center')
                        
                        # Adjust column widths
                        worksheet.column_dimensions['A'].width = 25  # gene
                        worksheet.column_dimensions['B'].width = 15  # term_type
                        worksheet.column_dimensions['C'].width = 30  # term
                    
                    print(f"✓ Created: {output_file_per_term} ({len(per_term_df)} rows)")
        else:
            print("No EggNOG v5 annotation files found")
    
    # Process v7 outputs
    if os.path.exists(eggnog_v7_dir):
        print("\nProcessing EggNOG v7 outputs...")
        tsv_gz_files = find_files(eggnog_v7_dir, '*.eggnog.tsv.gz')
        
        if tsv_gz_files:
            print(f"Found {len(tsv_gz_files)} EggNOG v7 annotation file(s)")
            for tsv_gz_file in tsv_gz_files:
                basename = os.path.basename(tsv_gz_file).replace('.eggnog.tsv.gz', '')
                
                parser = EggNOG7Parser(tsv_gz_file)
                parser.parse()
                
                # Save per-gene output (1 row = 1 gene)
                output_file_per_gene = os.path.join(output_dir, f"{basename}_eggnog_v7_per_gene.xlsx")
                parser.save_to_excel(output_file_per_gene)
                
                # Save per-term output (1 row = 1 GO or KEGG term)
                per_term_df = parser.create_per_term_output()
                if not per_term_df.empty:
                    output_file_per_term = os.path.join(output_dir, f"{basename}_eggnog_v7_per_term.xlsx")
                    with pd.ExcelWriter(output_file_per_term, engine='openpyxl') as writer:
                        per_term_df.to_excel(writer, index=False, sheet_name='Annotations')
                        
                        # Get workbook and worksheet
                        workbook = writer.book
                        worksheet = writer.sheets['Annotations']
                        
                        # Format header
                        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        header_font = Font(bold=True, color="FFFFFF")
                        
                        for cell in worksheet[1]:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal='center')
                        
                        # Adjust column widths
                        worksheet.column_dimensions['A'].width = 25  # gene
                        worksheet.column_dimensions['B'].width = 15  # term_type
                        worksheet.column_dimensions['C'].width = 30  # term
                        worksheet.column_dimensions['D'].width = 15  # score
                    
                    print(f"✓ Created: {output_file_per_term} ({len(per_term_df)} rows)")
        else:
            print("No EggNOG v7 annotation files found")
    
    if not os.path.exists(eggnog_v5_dir) and not os.path.exists(eggnog_v7_dir):
        print("Warning: No EggNOG directories found")


def process_fantasia(results_dir: str, output_dir: str):
    """Process all FANTASIA output files and create one Excel per model"""
    print("\n" + "="*60)
    print("Processing FANTASIA Results")
    print("="*60)
    
    fantasia_dir = os.path.join(results_dir, 'fantasia')
    if not os.path.exists(fantasia_dir):
        print(f"Warning: FANTASIA directory not found: {fantasia_dir}")
        return
    
    # Find all TSV files (FANTASIA outputs)
    tsv_files = find_files(fantasia_dir, '*.tsv')
    
    if not tsv_files:
        print("No FANTASIA TSV files found")
        return
    
    print(f"Found {len(tsv_files)} FANTASIA file(s)")
    
    # Define the models to process
    models = ['ESM_L0', 'Prot-T5_L0', 'Prost-T5_L0', 'Ankh3-Large_L0', 'ESM3c_L0']
    
    for tsv_file in tsv_files:
        basename = os.path.basename(tsv_file).replace('.tsv', '')
        
        # Process each model separately
        for model_suffix in models:
            model_name = FantasiaParser.MODEL_NAMES.get(model_suffix, model_suffix)
            
            parser = FantasiaParser(tsv_file, model_suffix=model_suffix)
            parser.parse()
            
            # Only save if there are results for this model
            if parser.results:
                # Save per-term output (1 row = 1 GO term)
                output_file_per_term = os.path.join(output_dir, f"{basename}_fantasia_{model_name}_per_term.xlsx")
                parser.save_to_excel(output_file_per_term)
                
                # Save per-gene output (1 row = 1 gene with grouped GO terms)
                per_gene_df = parser.create_per_gene_output()
                if not per_gene_df.empty:
                    output_file_per_gene = os.path.join(output_dir, f"{basename}_fantasia_{model_name}_per_gene.xlsx")
                    with pd.ExcelWriter(output_file_per_gene, engine='openpyxl') as writer:
                        per_gene_df.to_excel(writer, index=False, sheet_name='Annotations')
                        
                        # Get workbook and worksheet
                        workbook = writer.book
                        worksheet = writer.sheets['Annotations']
                        
                        # Format header
                        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        header_font = Font(bold=True, color="FFFFFF")
                        
                        for cell in worksheet[1]:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal='center')
                        
                        # Adjust column widths
                        worksheet.column_dimensions['A'].width = 25  # gene
                        worksheet.column_dimensions['B'].width = 80  # GO (grouped)
                    
                    print(f"✓ Created: {output_file_per_gene} ({len(per_gene_df)} rows)")
            else:
                print(f"  Skipping {model_name}: No annotations found")


def main():
    """Main function"""
    parser = argparse.ArgumentParser(
        description='Generate Excel files from functional annotation outputs',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Process all results in default directories
  %(prog)s

  # Specify custom results and output directories
  %(prog)s -r /path/to/results -o /path/to/output

  # Process only specific tools
  %(prog)s --kofamscan-only
  %(prog)s --interproscan-only
  %(prog)s --eggnog-only
  %(prog)s --fantasia-only
        """
    )
    
    parser.add_argument(
        '-r', '--results-dir',
        default='../../results',
        help='Results directory containing tool outputs (default: ../../results)'
    )
    
    parser.add_argument(
        '-o', '--output-dir',
        default='./excel_outputs',
        help='Output directory for Excel files (default: ./excel_outputs)'
    )
    
    parser.add_argument(
        '--kofamscan-only',
        action='store_true',
        help='Process only KofamScan results'
    )
    
    parser.add_argument(
        '--interproscan-only',
        action='store_true',
        help='Process only InterProScan results'
    )
    
    parser.add_argument(
        '--eggnog-only',
        action='store_true',
        help='Process only EggNOG-mapper results'
    )
    
    parser.add_argument(
        '--fantasia-only',
        action='store_true',
        help='Process only FANTASIA results'
    )
    
    args = parser.parse_args()
    
    # Resolve paths
    results_dir = os.path.abspath(args.results_dir)
    output_dir = os.path.abspath(args.output_dir)
    
    # Create output directory
    os.makedirs(output_dir, exist_ok=True)
    
    print("="*60)
    print("Functional Annotation Pipeline - Excel Output Generator")
    print("="*60)
    print(f"Results directory: {results_dir}")
    print(f"Output directory:  {output_dir}")
    print()
    
    if not os.path.exists(results_dir):
        print(f"Error: Results directory not found: {results_dir}")
        sys.exit(1)
    
    # Process tools based on flags
    process_all = not any([args.kofamscan_only, args.interproscan_only, args.eggnog_only, args.fantasia_only])
    
    if process_all or args.kofamscan_only:
        process_kofamscan(results_dir, output_dir)
    
    if process_all or args.interproscan_only:
        process_interproscan(results_dir, output_dir)
    
    if process_all or args.eggnog_only:
        process_eggnog(results_dir, output_dir)
    
    if process_all or args.fantasia_only:
        process_fantasia(results_dir, output_dir)
    
    print("\n" + "="*60)
    print("Excel Generation Complete")
    print("="*60)
    print(f"Output files saved to: {output_dir}")
    print()


if __name__ == '__main__':
    main()
