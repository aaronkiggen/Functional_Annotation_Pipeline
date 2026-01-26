#!/usr/bin/env python3
"""
Functional Annotation Pipeline - FANTASIA Results Filter

Description:
    This script performs post-processing filtering on FANTASIA results.
    It calculates 25th percentile thresholds for each model and applies
    filtering both to individual model Excel files and creates a consensus
    result across all models.
    
    Task A: Filter individual Excel files based on model-specific thresholds
    Task B: Create consensus filtering with majority vote (>=3/5 models)

Author: Aaron Kiggen
Repository: aaronkiggen/Functional_Annotation_Pipeline
"""

import argparse
import os
import sys
from pathlib import Path
from typing import Dict, Optional

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Error: Required packages not found.")
    print("Please install: pip install pandas openpyxl")
    sys.exit(1)


# Model name mapping: Excel file names -> CSV column names
MODEL_NAME_MAPPING = {
    'ESM-2': 'ESM_L0',
    'ProtT5': 'Prot-T5_L0',
    'ProstT5': 'Prost-T5_L0',
    'Ankh3-Large': 'Ankh3-Large_L0',
    'ESM3c': 'ESM3c_L0'
}


def parse_fasta_for_gene_mapping(fasta_path: str) -> Dict[str, str]:
    """
    Parse FASTA file to create protein-to-gene mapping.
    Extracts gene information from FASTA headers.
    
    Args:
        fasta_path: Path to protein FASTA file
        
    Returns:
        Dictionary mapping protein IDs to gene IDs
    """
    protein_to_gene = {}
    
    if not os.path.exists(fasta_path):
        print(f"Warning: FASTA file not found: {fasta_path}")
        return protein_to_gene
    
    with open(fasta_path, 'r') as f:
        for line in f:
            if line.startswith('>'):
                # Parse header - assuming format like >protein_id gene=gene_id
                header = line[1:].strip()
                parts = header.split()
                protein_id = parts[0]
                
                # Try to extract gene ID from header
                gene_id = protein_id  # Default to protein ID
                for part in parts:
                    if part.startswith('gene='):
                        gene_id = part.split('=')[1]
                        break
                
                protein_to_gene[protein_id] = gene_id
    
    return protein_to_gene


def calculate_thresholds(summary_df: pd.DataFrame) -> Dict[str, float]:
    """
    Calculate 25th percentile (Q1) thresholds for each model.
    
    Args:
        summary_df: DataFrame containing FANTASIA summary with model score columns
        
    Returns:
        Dictionary mapping model names to their Q1 threshold values
    """
    thresholds = {}
    
    # Model columns in the summary CSV
    model_columns = [
        'final_score_ESM_L0',
        'final_score_Prot-T5_L0',
        'final_score_Prost-T5_L0',
        'final_score_Ankh3-Large_L0',
        'final_score_ESM3c_L0'
    ]
    
    for col in model_columns:
        if col in summary_df.columns:
            # Get non-zero scores
            scores = summary_df[col].dropna()
            scores = pd.to_numeric(scores, errors='coerce')
            non_zero_scores = scores[scores > 0]
            
            if len(non_zero_scores) > 0:
                # Calculate 25th percentile (Q1)
                q1 = non_zero_scores.quantile(0.25)
                thresholds[col] = q1
                print(f"  {col}: Q1 = {q1:.4f} (from {len(non_zero_scores)} non-zero scores)")
            else:
                print(f"  {col}: No non-zero scores found")
                thresholds[col] = 0.0
        else:
            print(f"  Warning: Column {col} not found in summary")
    
    return thresholds


def filter_excel_file(excel_path: str, threshold: float, output_path: str):
    """
    Filter an Excel file based on final_score threshold.
    
    Args:
        excel_path: Path to input Excel file
        threshold: Minimum score threshold (Q1)
        output_path: Path to save filtered Excel file
    """
    # Load the Excel file
    df = pd.read_excel(excel_path, sheet_name='Annotations')
    
    original_count = len(df)
    
    # Filter rows where final_score >= threshold
    if 'final_score' in df.columns:
        # Convert to numeric, handling any non-numeric values
        df['final_score'] = pd.to_numeric(df['final_score'], errors='coerce')
        filtered_df = df[df['final_score'] >= threshold]
    else:
        print(f"  Warning: 'final_score' column not found in {excel_path}")
        filtered_df = df
    
    filtered_count = len(filtered_df)
    removed_count = original_count - filtered_count
    
    # Save filtered data
    if not filtered_df.empty:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            filtered_df.to_excel(writer, index=False, sheet_name='Annotations')
            
            # Format header
            workbook = writer.book
            worksheet = writer.sheets['Annotations']
            
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 60)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        print(f"  Filtered: {os.path.basename(excel_path)}")
        print(f"    Kept: {filtered_count}/{original_count} rows (removed {removed_count})")
    else:
        print(f"  Warning: All rows filtered out for {excel_path}")


def process_individual_excel_files(excel_dir: str, thresholds: Dict[str, float], output_dir: str):
    """
    Task A: Filter individual Excel files based on model-specific thresholds.
    
    Args:
        excel_dir: Directory containing Excel files
        thresholds: Dictionary of model thresholds
        output_dir: Directory to save filtered outputs
    """
    print("\n" + "="*60)
    print("Task A: Filtering Individual Excel Files")
    print("="*60)
    
    # Find all FANTASIA per-term Excel files
    excel_files = list(Path(excel_dir).glob('*_fantasia_*_per_term.xlsx'))
    
    if not excel_files:
        print("No FANTASIA per-term Excel files found")
        return
    
    print(f"Found {len(excel_files)} Excel file(s)")
    
    for excel_file in excel_files:
        excel_path = str(excel_file)
        filename = excel_file.name
        
        # Determine which model this file corresponds to
        model_name = None
        for excel_model in MODEL_NAME_MAPPING.keys():
            if f'_fantasia_{excel_model}_' in filename:
                model_name = excel_model
                break
        
        if not model_name:
            print(f"  Warning: Could not determine model for {filename}")
            continue
        
        # Get corresponding threshold
        csv_model = MODEL_NAME_MAPPING[model_name]
        threshold_col = f'final_score_{csv_model}'
        
        if threshold_col not in thresholds:
            print(f"  Warning: No threshold found for {threshold_col}")
            continue
        
        threshold = thresholds[threshold_col]
        
        # Create output filename
        output_filename = filename.replace('_per_term.xlsx', '_per_term_filtered.xlsx')
        output_path = os.path.join(output_dir, output_filename)
        
        # Filter the Excel file
        filter_excel_file(excel_path, threshold, output_path)


def create_consensus_filtering(summary_path: str, thresholds: Dict[str, float], 
                               output_dir: str, fasta_path: Optional[str] = None):
    """
    Task B: Create consensus filtering with majority vote (>=3/5 models).
    
    Args:
        summary_path: Path to FANTASIA summary CSV
        thresholds: Dictionary of model thresholds
        output_dir: Directory to save consensus results
        fasta_path: Optional path to FASTA file for gene mapping
    """
    print("\n" + "="*60)
    print("Task B: Consensus Filtering (Majority Vote >= 3/5)")
    print("="*60)
    
    # Load summary CSV
    summary_df = pd.read_csv(summary_path, sep='\t')
    print(f"Loaded summary with {len(summary_df)} rows")
    
    # Model score columns
    model_columns = [
        'final_score_ESM_L0',
        'final_score_Prot-T5_L0',
        'final_score_Prost-T5_L0',
        'final_score_Ankh3-Large_L0',
        'final_score_ESM3c_L0'
    ]
    
    # Apply thresholds and calculate consensus
    consensus_votes = []
    
    for idx, row in summary_df.iterrows():
        vote_count = 0
        
        for col in model_columns:
            if col in summary_df.columns and col in thresholds:
                score = pd.to_numeric(row[col], errors='coerce')
                if pd.notna(score) and score >= thresholds[col]:
                    vote_count += 1
        
        consensus_votes.append(vote_count)
    
    summary_df['consensus_vote'] = consensus_votes
    
    # Filter for majority vote (>= 3)
    consensus_df = summary_df[summary_df['consensus_vote'] >= 3].copy()
    
    print(f"Entries passing consensus filter: {len(consensus_df)}/{len(summary_df)}")
    print(f"  Vote distribution:")
    vote_dist = summary_df['consensus_vote'].value_counts().sort_index()
    for vote, count in vote_dist.items():
        marker = "✓" if vote >= 3 else " "
        print(f"    {marker} {vote} votes: {count} entries")
    
    # Apply protein-to-gene mapping if FASTA provided
    if fasta_path and os.path.exists(fasta_path):
        print("\nApplying protein-to-gene mapping...")
        protein_to_gene = parse_fasta_for_gene_mapping(fasta_path)
        
        if protein_to_gene and 'accession' in consensus_df.columns:
            consensus_df['gene_id'] = consensus_df['accession'].map(protein_to_gene)
            # If mapping fails, use accession as gene_id
            consensus_df['gene_id'] = consensus_df['gene_id'].fillna(consensus_df['accession'])
            print(f"  Mapped {len(protein_to_gene)} proteins to genes")
    
    # Save consensus result as CSV
    output_csv = os.path.join(output_dir, 'fantasia_consensus_majority.csv')
    consensus_df.to_csv(output_csv, index=False)
    print(f"\n✓ Saved consensus result: {output_csv}")
    
    # Also save as Excel with formatting
    output_xlsx = os.path.join(output_dir, 'fantasia_consensus_majority.xlsx')
    with pd.ExcelWriter(output_xlsx, engine='openpyxl') as writer:
        consensus_df.to_excel(writer, index=False, sheet_name='Consensus')
        
        # Format header
        workbook = writer.book
        worksheet = writer.sheets['Consensus']
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    print(f"✓ Saved consensus result: {output_xlsx}")


def main():
    """Main function to parse arguments and run filtering"""
    parser = argparse.ArgumentParser(
        description='Filter FANTASIA results using 25th percentile thresholds',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Basic usage
  python filter_fantasia_results.py --summary results.tsv --excel-dir excel_outputs/
  
  # With custom output directory
  python filter_fantasia_results.py --summary results.tsv --excel-dir excel_outputs/ --output-dir filtered/
  
  # With FASTA for gene mapping
  python filter_fantasia_results.py --summary results.tsv --excel-dir excel_outputs/ --fasta proteins.faa
        """
    )
    
    parser.add_argument(
        '--summary',
        required=True,
        help='Path to FANTASIA summary CSV/TSV file'
    )
    
    parser.add_argument(
        '--excel-dir',
        required=True,
        help='Directory containing Excel files generated by create_excel_outputs.py'
    )
    
    parser.add_argument(
        '--fasta',
        help='Path to protein FASTA file (optional, for gene mapping)'
    )
    
    parser.add_argument(
        '--output-dir',
        help='Directory to save filtered outputs (default: same as excel-dir)'
    )
    
    args = parser.parse_args()
    
    # Validate inputs
    if not os.path.exists(args.summary):
        print(f"Error: Summary file not found: {args.summary}")
        sys.exit(1)
    
    if not os.path.exists(args.excel_dir):
        print(f"Error: Excel directory not found: {args.excel_dir}")
        sys.exit(1)
    
    # Set output directory
    output_dir = args.output_dir if args.output_dir else args.excel_dir
    os.makedirs(output_dir, exist_ok=True)
    
    print("="*60)
    print("FANTASIA Results Filter")
    print("="*60)
    print(f"Summary file: {args.summary}")
    print(f"Excel directory: {args.excel_dir}")
    print(f"Output directory: {output_dir}")
    if args.fasta:
        print(f"FASTA file: {args.fasta}")
    
    # Load summary and calculate thresholds
    print("\n" + "="*60)
    print("Calculating 25th Percentile Thresholds")
    print("="*60)
    
    summary_df = pd.read_csv(args.summary, sep='\t')
    thresholds = calculate_thresholds(summary_df)
    
    if not thresholds:
        print("Error: No thresholds calculated")
        sys.exit(1)
    
    # Task A: Filter individual Excel files
    process_individual_excel_files(args.excel_dir, thresholds, output_dir)
    
    # Task B: Create consensus filtering
    create_consensus_filtering(args.summary, thresholds, output_dir, args.fasta)
    
    print("\n" + "="*60)
    print("Filtering Complete!")
    print("="*60)


if __name__ == '__main__':
    main()
